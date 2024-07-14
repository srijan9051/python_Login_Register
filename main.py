from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pathlib
import mysql.connector

background = "#062830"
framebg = "#EDEDED"
framefg = "#062830"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    headers = ["Registration No.", "Name", "Class", "Gender", "DOB", "Date of Registration", "Religion", "Skill", "Father Name", "Mother Name", "Father's Occupation", "Mother's Occupation"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    file.save('Student_data.xlsx')

# gender
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

def search():
    text=Search.get()
    clear()
    SaveButton.config(state='disable')
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid Registration Number")
    
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=='Female':
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img=(Image.open("Uploaded_Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1=gender
    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    try:
        mydb = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Srijan@2000',
            database='StudentRegistration'
        )
        mycursor = mydb.cursor()

        # Fetch the row based on registration number R1
        sql_select_query = "SELECT * FROM student WHERE Reg_no = %s"
        mycursor.execute(sql_select_query, (R1,))
        record = mycursor.fetchone()

        if record is None:
            messagebox.showerror("Error", "Registration number not found")
            mycursor.close()
            mydb.close()
            return

        # Update the row in the database table student
        sql_update_query = """
            UPDATE student 
            SET Name = %s, Class = %s, Gender = %s, DOB = %s, registration_date = %s, 
                Religion = %s, Skill = %s, Father_Name = %s, Mother_Name = %s, 
                Father_Occupation = %s, Mother_Occupation = %s 
            WHERE Reg_no = %s
        """
        val = (N1, C1, G1, D2, D1, Rel, S1, fathername, mothername, F1, M1, R1)
        mycursor.execute(sql_update_query, val)
        mydb.commit()
        mycursor.close()
        mydb.close()
    except mysql.connector.Error as err:
        messagebox.showerror("Failed", f"Error: {err}")
        return

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            reg_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    #sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Rel)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=fathername)
    sheet.cell(column=10,row=int(reg_number),value=mothername)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Uploaded_Student Images/"+str(R1)+".jpg")
    except:
        messagebox.showerror("Error","Could not upload Image")
    messagebox.showinfo("Success","Updated Successfully")

    
    clear()



def Exit():
    root.destroy()

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file",
                                          filetype=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set(1)




def clear():
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")

    registration_no()
    SaveButton.config(state='normal')
    img1 = PhotoImage(file="Image/upload photo.png")
    lbl.config(image=img1)
    lbl.image = img1

    img = ""

def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Please select the gender")
        return

    D2 = DOB.get()
    D1 = Date.get()  # Registration Date
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "Some entry missing...")
    else:
        try:
            mydb = mysql.connector.connect(
                host='localhost',
                user='root',
                password='Srijan@2000',
                database='StudentRegistration'
            )
            mycursor = mydb.cursor()

            # Check if the table exists, if not, create it
            mycursor.execute("""
                CREATE TABLE IF NOT EXISTS student (
                    Reg_no INT AUTO_INCREMENT PRIMARY KEY,
                    Name VARCHAR(50),
                    Class VARCHAR(100),
                    Gender VARCHAR(50),
                    DOB VARCHAR(50),
                    registration_date VARCHAR(50),
                    Religion VARCHAR(50),
                    Skill VARCHAR(100),
                    Father_Name VARCHAR(50),
                    Mother_Name VARCHAR(50),
                    Father_Occupation VARCHAR(100),
                    Mother_Occupation VARCHAR(100)
                )
            """)

            # Insert the student details into the table
            sql = """
                INSERT INTO student 
                (Reg_no, Name, Class, Gender, DOB, registration_date, Religion, Skill, Father_Name, Mother_Name, Father_Occupation, Mother_Occupation) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            val = (R1, N1, C1, G1, D2, D1, Rel, S1, fathername, mothername, F1, M1)
            mycursor.execute(sql, val)
            mydb.commit()
            mycursor.close()
            mydb.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Failed", f"Error: {err}")
            return

        # Insert into the Excel sheet
        try:
            file = openpyxl.load_workbook('Student_data.xlsx')
            sheet = file.active
            row = sheet.max_row + 1

            data = [R1, N1, C1, G1, D2, D1, Rel, S1, fathername, mothername, F1, M1]
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col_num, value=value)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            file.save('Student_data.xlsx')
            try:
                img.save("Uploaded_Student Images/" + str(R1) + ".jpg")
            except:
                messagebox.showinfo("Info", "Picture not available!!")
            messagebox.showinfo("Success", "All details uploaded successfully")
            clear()
            registration_no()
        except Exception as e:
            messagebox.showerror("Failed", f"Upload Failed: {e}")


# top frames
Label(root, text="Email: srijanbanerjee18@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=3, bg="#C36464", fg="#fff", font='Arial 20 bold').pack(side=TOP, fill=X)

# search box
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=860, y=85)
imageicon3 = PhotoImage(file="Image/search.png")
srch = Button(root, text="Search", compound=LEFT, image=imageicon3, cursor="hand2", width=123, bg='#68ddfa', font="arial 13 bold",command=search)
srch.place(x=1100, y=85)
imageicon4 = PhotoImage(file="Image/Layer 4.png")
update_button = Button(root, image=imageicon4, bg="#c36464",cursor="hand2",command=Update)
update_label=Label(root,text="UPDATE",font="arial 13",fg=framebg,bg=background).place(x=110,y=105)
update_button.place(x=110, y=64)

# Register Button
Label(root, text="Registration No.", font="arial 13", fg=framebg, bg=background).place(x=30, y=170)
Label(root, text="Date", font="arial 13", fg=framebg, bg=background).place(x=5000, y=170)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=170)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=170)
Date.set(d1)

# Student Details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=210)

Label(obj, text="Full Name", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font="arial 10")
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="arial 10", width=17, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

# Parent Details
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=480)

Label(obj2, text="Father's Name", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="arial 10")
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name = StringVar()
m_entry = Entry(obj2, textvariable=M_Name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="arial 10")
MO_entry.place(x=630, y=100)

# image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=170)

img = PhotoImage(file="Image/upload photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", cursor="hand2", command=showimage).place(x=1000, y=370)

SaveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", cursor="hand2", command=Save)
SaveButton.place(x=1000, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", cursor="hand2", command=clear).place(x=1000, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", cursor="hand2", command=Exit).place(x=1000, y=610)

root.mainloop()
